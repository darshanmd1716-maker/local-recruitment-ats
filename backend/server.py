from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient

import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import json
import re
import asyncio
import shutil

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import aiofiles

# --- Resume text extraction libs ---
from pypdf import PdfReader
from docx import Document

# --- Gemini ---
import google.generativeai as genai


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# MongoDB connection
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Ensure upload directories exist
UPLOAD_DIR = ROOT_DIR / "uploads"
RECRUITMENT_DIR = ROOT_DIR / "Recruitment"
EXPORT_DIR = ROOT_DIR / "exports"
UPLOAD_DIR.mkdir(exist_ok=True)
RECRUITMENT_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


# ===================== Gemini helpers =====================

def configure_gemini() -> None:
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY is missing in environment variables")
    genai.configure(api_key=api_key)


async def gemini_text(prompt: str, model: str = "gemini-1.5-flash") -> str:
    """
    google.generativeai SDK is sync; run in thread to avoid blocking.
    """
    configure_gemini()

    def _run() -> str:
        m = genai.GenerativeModel(model)
        resp = m.generate_content(prompt)
        return (resp.text or "").strip()

    return await asyncio.to_thread(_run)


def safe_json_from_text(text: str) -> Optional[Dict[str, Any]]:
    """
    Extract first JSON object from a model response.
    """
    match = re.search(r"\{[\s\S]*\}", text)
    if not match:
        return None
    try:
        return json.loads(match.group())
    except Exception:
        return None


# ===================== Models =====================

class JobDescription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    raw_text: str
    required_skills: List[str] = []
    experience_required: Optional[str] = None
    location: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class JobDescriptionCreate(BaseModel):
    title: str
    raw_text: str


class Candidate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    job_id: str
    name: str
    mobile: Optional[str] = None
    email: Optional[str] = None
    skills: List[str] = []
    experience: Optional[str] = None
    current_role: Optional[str] = None
    match_percentage: float = 0.0
    category: str = "Rejected_Future"  # Shortlisted, Hold, Rejected_Future
    original_filename: str = ""
    current_ctc: Optional[str] = None
    expected_ctc: Optional[str] = None
    notice_period: Optional[str] = None
    negotiable: Optional[str] = None
    candidate_response: Optional[str] = "Pending"
    remarks: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ProcessingResult(BaseModel):
    job_id: str
    total_processed: int
    shortlisted: int
    hold: int
    rejected_future: int
    top_candidates: List[Dict[str, Any]]
    folder_created: bool
    excel_path: Optional[str] = None
    duplicates_found: List[Dict[str, Any]] = []


class CompareRequest(BaseModel):
    candidate_ids: List[str]


# ===================== Helper Functions =====================

def normalize_phone(phone: Optional[str]) -> Optional[str]:
    if not phone:
        return None
    digits = re.sub(r"\D", "", phone)
    return digits[-10:] if len(digits) >= 10 else (digits if digits else None)


def normalize_email(email: Optional[str]) -> Optional[str]:
    if not email:
        return None
    return email.lower().strip()


async def check_duplicate_candidate(
    email: Optional[str],
    mobile: Optional[str],
    exclude_job_id: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    normalized_email = normalize_email(email)
    normalized_mobile = normalize_phone(mobile)

    if not normalized_email and not normalized_mobile:
        return None

    or_conditions = []
    if normalized_email:
        or_conditions.append({"email": {"$regex": f"^{re.escape(normalized_email)}$", "$options": "i"}})
    if normalized_mobile:
        or_conditions.append({"mobile": {"$regex": f"{normalized_mobile}$"}})

    if not or_conditions:
        return None

    query: Dict[str, Any] = {"$or": or_conditions}
    if exclude_job_id:
        query["job_id"] = {"$ne": exclude_job_id}

    existing = await db.candidates.find_one(query, {"_id": 0})
    if existing:
        existing_email = normalize_email(existing.get("email"))
        existing_mobile = normalize_phone(existing.get("mobile"))

        match_type = []
        if normalized_email and existing_email and normalized_email == existing_email:
            match_type.append("email")
        if normalized_mobile and existing_mobile and normalized_mobile == existing_mobile:
            match_type.append("mobile")

        existing["match_type"] = " & ".join(match_type) if match_type else "unknown"
        return existing

    return None


# ===================== Resume text extraction =====================

def extract_text_from_pdf(path: str) -> str:
    try:
        reader = PdfReader(path)
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(parts).strip()
    except Exception as e:
        logger.error(f"PDF text extraction failed: {e}")
        return ""


def extract_text_from_docx(path: str) -> str:
    try:
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs).strip()
    except Exception as e:
        logger.error(f"DOCX text extraction failed: {e}")
        return ""


def extract_text_from_txt_bytes(content: bytes) -> str:
    try:
        return content.decode("utf-8", errors="ignore").strip()
    except Exception:
        return ""


def extract_resume_text(file_path: str, content: bytes, filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    if ext == ".docx":
        return extract_text_from_docx(file_path)
    if ext == ".txt":
        return extract_text_from_txt_bytes(content)

    # .doc is tricky without extra libs; treat as empty for now.
    # You can convert .doc to .docx when downloading from Naukri if possible.
    return ""


# ===================== AI Functions (Gemini) =====================

async def parse_jd_with_ai(jd_text: str) -> Dict[str, Any]:
    try:
        prompt = f"""Extract structured details from the Job Description and return ONLY valid JSON:

Schema:
{{
  "title": "Job title",
  "required_skills": ["skill1", "skill2"],
  "experience_required": "X years / range / text",
  "location": "Location or Remote"
}}

Job Description:
{jd_text}

Rules:
- Return only JSON. No markdown, no explanation.
- If unsure, use null or empty list.
"""
        response = await gemini_text(prompt)
        data = safe_json_from_text(response)
        if not data:
            return {"title": "Unknown", "required_skills": [], "experience_required": None, "location": None}

        # Normalize fields
        return {
            "title": data.get("title") or "Unknown",
            "required_skills": data.get("required_skills") or [],
            "experience_required": data.get("experience_required"),
            "location": data.get("location"),
        }
    except Exception as e:
        logger.error(f"Error parsing JD with Gemini: {e}")
        return {"title": "Unknown", "required_skills": [], "experience_required": None, "location": None}


async def parse_resume_with_ai(file_path: str, file_content: bytes, filename: str) -> Dict[str, Any]:
    """
    Extract candidate data from resume text (PDF/DOCX/TXT) using Gemini.
    """
    try:
        resume_text = extract_resume_text(file_path, file_content, filename)
        if not resume_text:
            # still store file + candidate record; but fields will be unknown
            return {"name": "Unknown", "mobile": None, "email": None, "skills": [], "experience": None, "current_role": None}

        # Keep prompt size reasonable
        clipped = resume_text[:12000]

        prompt = f"""You are an expert IT recruiter. Extract candidate information from the RESUME TEXT below.
Return ONLY valid JSON exactly with this schema:

{{
  "name": "Full name of candidate",
  "mobile": "Phone number (string) or null",
  "email": "Email address or null",
  "skills": ["skill1","skill2"],
  "experience": "Total years of experience as STRING (e.g., '5 years', '3+ years', '2-4 years') or null",
  "current_role": "Current/last job title or null"
}}

Rules:
- Return only JSON. No markdown, no explanation.
- If not found, use null/empty list.
- experience MUST be a string if present.

RESUME TEXT:
{clipped}
"""
        response = await gemini_text(prompt)
        data = safe_json_from_text(response)
        if not data:
            return {"name": "Unknown", "mobile": None, "email": None, "skills": [], "experience": None, "current_role": None}

        # Ensure experience string
        exp = data.get("experience")
        if exp is not None and not isinstance(exp, str):
            exp = f"{exp} years"

        return {
            "name": data.get("name") or "Unknown",
            "mobile": data.get("mobile"),
            "email": data.get("email"),
            "skills": data.get("skills") or [],
            "experience": exp,
            "current_role": data.get("current_role"),
        }
    except Exception as e:
        logger.error(f"Error parsing resume with Gemini: {e}")
        return {"name": "Unknown", "mobile": None, "email": None, "skills": [], "experience": None, "current_role": None}


async def calculate_match_score(jd_skills: List[str], jd_text: str, candidate_data: Dict[str, Any]) -> float:
    """
    Gemini semantic scoring + keyword fallback.
    """
    try:
        prompt = f"""As an IT recruiter, score candidate-job match from 0 to 100.

Job:
- Required skills: {", ".join(jd_skills)}
- JD (short): {jd_text[:1000]}

Candidate:
- Name: {candidate_data.get("name", "Unknown")}
- Skills: {", ".join(candidate_data.get("skills", []))}
- Experience: {candidate_data.get("experience", "Unknown")}
- Current role: {candidate_data.get("current_role", "Unknown")}

Return ONLY a number between 0 and 100. No other text.
"""
        response = await gemini_text(prompt)
        nums = re.findall(r"\d+(?:\.\d+)?", response)
        if nums:
            score = float(nums[0])
            return float(min(100, max(0, score)))
        return 50.0
    except Exception as e:
        logger.error(f"Error calculating match score with Gemini: {e}")
        # Keyword fallback
        if not jd_skills:
            return 50.0
        candidate_skills = set(s.lower() for s in candidate_data.get("skills", []))
        jd_skills_lower = set(s.lower() for s in jd_skills)
        if not jd_skills_lower:
            return 50.0
        overlap = len(candidate_skills.intersection(jd_skills_lower))
        return min(100.0, (overlap / len(jd_skills_lower)) * 100.0)


def categorize_candidate(match_percentage: float) -> str:
    if match_percentage >= 75:
        return "Shortlisted"
    elif match_percentage >= 50:
        return "Hold"
    else:
        return "Rejected_Future"


def create_recruitment_folders(job_title: str) -> Dict[str, Path]:
    clean_title = re.sub(r"[^\w\s-]", "", job_title).strip().replace(" ", "_")
    job_folder = RECRUITMENT_DIR / clean_title

    folders = {
        "root": job_folder,
        "shortlisted": job_folder / "Shortlisted",
        "hold": job_folder / "Hold",
        "rejected_future": job_folder / "Rejected_Future",
    }

    for folder in folders.values():
        folder.mkdir(parents=True, exist_ok=True)

    return folders


async def generate_excel_tracker(job_id: str, job_title: str) -> str:
    candidates = await db.candidates.find({"job_id": job_id}, {"_id": 0}).to_list(1000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = job_title[:31]

    headers = [
        "Candidate Name", "Mobile", "Email", "Skills", "Experience",
        "Match Percentage", "Current CTC", "Expected CTC",
        "Notice Period / LWD", "Negotiable", "Candidate Response", "Remarks",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    category_colors = {
        "Shortlisted": "22C55E",
        "Hold": "EAB308",
        "Rejected_Future": "EF4444",
    }

    for row, candidate in enumerate(candidates, 2):
        data = [
            candidate.get("name", ""),
            candidate.get("mobile", ""),
            candidate.get("email", ""),
            ", ".join(candidate.get("skills", [])),
            candidate.get("experience", ""),
            f"{candidate.get('match_percentage', 0):.1f}%",
            candidate.get("current_ctc", ""),
            candidate.get("expected_ctc", ""),
            candidate.get("notice_period", ""),
            candidate.get("negotiable", ""),
            candidate.get("candidate_response", "Pending"),
            candidate.get("remarks", ""),
        ]

        category = candidate.get("category", "Rejected_Future")
        row_fill = PatternFill(
            start_color=category_colors.get(category, "27272A"),
            end_color=category_colors.get(category, "27272A"),
            fill_type="solid",
        )

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 6:
                cell.fill = row_fill
                cell.font = Font(bold=True, color="FFFFFF")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    filename = f"Recruitment_Tracker_{job_title[:20]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORT_DIR / filename
    wb.save(filepath)

    return str(filepath)


# ===================== API Routes =====================

@api_router.get("/")
async def root():
    return {"message": "Local Recruitment ATS API"}


@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}


@api_router.post("/jobs", response_model=JobDescription)
async def create_job(job_input: JobDescriptionCreate):
    parsed_data = await parse_jd_with_ai(job_input.raw_text)

    job = JobDescription(
        title=job_input.title or parsed_data.get("title", "Unknown"),
        raw_text=job_input.raw_text,
        required_skills=parsed_data.get("required_skills", []),
        experience_required=parsed_data.get("experience_required"),
        location=parsed_data.get("location"),
    )

    doc = job.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.jobs.insert_one(doc)

    create_recruitment_folders(job.title)
    return job


@api_router.get("/jobs", response_model=List[JobDescription])
async def get_jobs():
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(100)
    for job in jobs:
        if isinstance(job.get("created_at"), str):
            job["created_at"] = datetime.fromisoformat(job["created_at"])
    return jobs


@api_router.get("/jobs/{job_id}", response_model=JobDescription)
async def get_job(job_id: str):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if isinstance(job.get("created_at"), str):
        job["created_at"] = datetime.fromisoformat(job["created_at"])
    return job


@api_router.delete("/jobs/{job_id}")
async def delete_job(job_id: str):
    result = await db.jobs.delete_one({"id": job_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")
    await db.candidates.delete_many({"job_id": job_id})
    return {"message": "Job deleted successfully"}


@api_router.post("/process-resumes/{job_id}", response_model=ProcessingResult)
async def process_resumes(job_id: str, files: List[UploadFile] = File(...)):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    folders = create_recruitment_folders(job["title"])

    candidates_processed: List[Dict[str, Any]] = []
    shortlisted_count = 0
    hold_count = 0
    rejected_count = 0
    duplicates_found: List[Dict[str, Any]] = []

    for file in files:
        ext = Path(file.filename).suffix.lower()
        if ext not in [".pdf", ".doc", ".docx", ".txt"]:
            continue

        temp_path = UPLOAD_DIR / f"{uuid.uuid4()}_{file.filename}"
        content = await file.read()

        async with aiofiles.open(temp_path, "wb") as f:
            await f.write(content)

        try:
            parsed_data = await parse_resume_with_ai(str(temp_path), content, file.filename)

            duplicate = await check_duplicate_candidate(
                parsed_data.get("email"),
                parsed_data.get("mobile"),
            )

            if duplicate:
                existing_job = await db.jobs.find_one({"id": duplicate.get("job_id")}, {"_id": 0})
                duplicates_found.append({
                    "new_name": parsed_data.get("name", "Unknown"),
                    "new_email": parsed_data.get("email"),
                    "new_mobile": parsed_data.get("mobile"),
                    "existing_name": duplicate.get("name"),
                    "existing_email": duplicate.get("email"),
                    "existing_mobile": duplicate.get("mobile"),
                    "existing_job": existing_job.get("title") if existing_job else "Unknown",
                    "existing_category": duplicate.get("category"),
                    "existing_match": duplicate.get("match_percentage"),
                    "match_type": duplicate.get("match_type", "unknown"),
                    "filename": file.filename,
                })

            match_score = await calculate_match_score(
                job.get("required_skills", []),
                job.get("raw_text", ""),
                parsed_data,
            )

            category = categorize_candidate(match_score)

            candidate = Candidate(
                job_id=job_id,
                name=parsed_data.get("name", "Unknown"),
                mobile=parsed_data.get("mobile"),
                email=parsed_data.get("email"),
                skills=parsed_data.get("skills", []),
                experience=parsed_data.get("experience"),
                current_role=parsed_data.get("current_role"),
                match_percentage=match_score,
                category=category,
                original_filename=file.filename,
                remarks=f"DUPLICATE: Found {duplicate.get('match_type')} match with existing candidate" if duplicate else None,
            )

            doc = candidate.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            doc["is_duplicate"] = True if duplicate else False
            await db.candidates.insert_one(doc)

            category_folder = {
                "Shortlisted": folders["shortlisted"],
                "Hold": folders["hold"],
                "Rejected_Future": folders["rejected_future"],
            }
            dest_path = category_folder[category] / file.filename
            shutil.copy2(temp_path, dest_path)

            candidates_processed.append(candidate.model_dump())

            if category == "Shortlisted":
                shortlisted_count += 1
            elif category == "Hold":
                hold_count += 1
            else:
                rejected_count += 1

        finally:
            if temp_path.exists():
                temp_path.unlink()

    excel_path = await generate_excel_tracker(job_id, job["title"])
    top_candidates = sorted(candidates_processed, key=lambda x: x["match_percentage"], reverse=True)[:5]

    return ProcessingResult(
        job_id=job_id,
        total_processed=len(candidates_processed),
        shortlisted=shortlisted_count,
        hold=hold_count,
        rejected_future=rejected_count,
        top_candidates=top_candidates,
        folder_created=True,
        excel_path=Path(excel_path).name if excel_path else None,
        duplicates_found=duplicates_found,
    )


@api_router.get("/candidates/{job_id}", response_model=List[Candidate])
async def get_candidates(job_id: str, category: Optional[str] = None):
    query: Dict[str, Any] = {"job_id": job_id}
    if category:
        query["category"] = category

    candidates = await db.candidates.find(query, {"_id": 0}).to_list(1000)
    for candidate in candidates:
        if isinstance(candidate.get("created_at"), str):
            candidate["created_at"] = datetime.fromisoformat(candidate["created_at"])
    return candidates


@api_router.put("/candidates/{candidate_id}")
async def update_candidate(candidate_id: str, updates: Dict[str, Any]):
    allowed_fields = ["current_ctc", "expected_ctc", "notice_period", "negotiable", "candidate_response", "remarks"]
    filtered_updates = {k: v for k, v in updates.items() if k in allowed_fields}

    if not filtered_updates:
        raise HTTPException(status_code=400, detail="No valid fields to update")

    result = await db.candidates.update_one(
        {"id": candidate_id},
        {"$set": filtered_updates},
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Candidate not found")

    return {"message": "Candidate updated successfully"}


@api_router.delete("/candidates/{candidate_id}")
async def delete_candidate(candidate_id: str):
    result = await db.candidates.delete_one({"id": candidate_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Candidate not found")
    return {"message": "Candidate deleted successfully"}


@api_router.get("/export/{job_id}")
async def export_excel(job_id: str):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    excel_path = await generate_excel_tracker(job_id, job["title"])

    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=Path(excel_path).name,
    )


@api_router.get("/stats")
async def get_stats():
    total_jobs = await db.jobs.count_documents({})
    total_candidates = await db.candidates.count_documents({})
    shortlisted = await db.candidates.count_documents({"category": "Shortlisted"})
    hold = await db.candidates.count_documents({"category": "Hold"})
    rejected = await db.candidates.count_documents({"category": "Rejected_Future"})

    return {
        "total_jobs": total_jobs,
        "total_candidates": total_candidates,
        "shortlisted": shortlisted,
        "hold": hold,
        "rejected_future": rejected,
    }


@api_router.post("/compare-candidates")
async def compare_candidates(request: CompareRequest):
    if len(request.candidate_ids) < 2:
        raise HTTPException(status_code=400, detail="At least 2 candidates required for comparison")
    if len(request.candidate_ids) > 5:
        raise HTTPException(status_code=400, detail="Maximum 5 candidates can be compared at once")

    candidates = []
    for cid in request.candidate_ids:
        candidate = await db.candidates.find_one({"id": cid}, {"_id": 0})
        if candidate:
            if isinstance(candidate.get("created_at"), str):
                candidate["created_at"] = datetime.fromisoformat(candidate["created_at"])
            job = await db.jobs.find_one({"id": candidate.get("job_id")}, {"_id": 0})
            candidate["job_title"] = job.get("title") if job else "Unknown"
            candidates.append(candidate)

    if len(candidates) < 2:
        raise HTTPException(status_code=404, detail="Not enough valid candidates found")

    all_skills = set()
    for c in candidates:
        all_skills.update(s.lower() for s in c.get("skills", []))

    for c in candidates:
        candidate_skills = set(s.lower() for s in c.get("skills", []))
        c["skill_coverage"] = round((len(candidate_skills) / len(all_skills)) * 100, 1) if all_skills else 0
        c["unique_skills"] = list(
            candidate_skills - set().union(*[
                set(s.lower() for s in other.get("skills", []))
                for other in candidates if other["id"] != c["id"]
            ])
        )

    common_skills = set(s.lower() for s in candidates[0].get("skills", []))
    for c in candidates[1:]:
        common_skills &= set(s.lower() for s in c.get("skills", []))

    return {
        "candidates": candidates,
        "comparison_metrics": {
            "total_unique_skills": len(all_skills),
            "common_skills": list(common_skills),
            "common_skills_count": len(common_skills),
        },
    }


@api_router.get("/check-duplicate")
async def check_duplicate(email: Optional[str] = None, mobile: Optional[str] = None):
    if not email and not mobile:
        raise HTTPException(status_code=400, detail="Email or mobile required")

    duplicate = await check_duplicate_candidate(email, mobile)
    if duplicate:
        job = await db.jobs.find_one({"id": duplicate.get("job_id")}, {"_id": 0})
        duplicate["job_title"] = job.get("title") if job else "Unknown"
        return {"is_duplicate": True, "existing_candidate": duplicate}

    return {"is_duplicate": False, "existing_candidate": None}


@api_router.get("/duplicates/{job_id}")
async def get_duplicates(job_id: str):
    candidates = await db.candidates.find(
        {"job_id": job_id, "is_duplicate": True},
        {"_id": 0},
    ).to_list(100)

    for candidate in candidates:
        if isinstance(candidate.get("created_at"), str):
            candidate["created_at"] = datetime.fromisoformat(candidate["created_at"])

    return candidates


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()